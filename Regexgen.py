import openpyxl
from tkinter import *


def regCreatorT(daterange,FIP,Sub,VLN,wsNet,wsDate,subsplitNet):
    FOct = subsplitNet.split('.')[0]
    StOct = subsplitNet.split('.')[1]
    TOct = subsplitNet.split('.')[2]
    LOct=subsplitNet.split('.')[3]
    IPRegx=FOct+"\."+StOct+"\."+TOct+"\."
    LastRegx=""
    print("ankit")
    print(str(LOct))
    for i in range(1, wsNet.max_row + 1):
        print(str(wsNet.cell(row=i, column=1).value))
        if wsNet.cell(row=i, column=1).value is not None:
            print("inside row "+str(wsNet.cell(row=i, column=1).value))
            if str(wsNet.cell(row=i, column=1).value) == str(LOct):
                for j in range(1, wsNet.max_column  + 1):
                    print("inside column " + str(wsNet.cell(row=i, column=1).value))
                    if wsNet.cell(row=i, column=j).value is not None:
                        if j!=1:
                            LastRegx=LastRegx+IPRegx+str(wsNet.cell(row=i, column=j).value)+"\s|"
    for i in range(1, wsDate.max_row + 1):
        if wsDate.cell(row=i, column=3).value is not None:
            if str(wsDate.cell(row=i, column=3).value) == str(daterange):
                LastDateRegex=wsDate.cell(row=i, column=4).value

    dateRegexFinalT="zcat /data/log/2021-"+str(LastDateRegex)+"-[0-2][0-9]*/"+str(FIP)+"-traffic* | grep -E "
    IPFinalRegex="'"+LastRegx[:-1]+"'"
    endStringT = '''| awk -F "=" '{print $7 }'  | awk 'NF{NF-=1};1' | awk -F"ankitdeshwal" '{a[$1]++}END{for(i in a) {print i, a[i]}}' >> ''' + str(VLN)+"-Toptalker.txt"
    toptalker=dateRegexFinalT+IPFinalRegex+endStringT
    dateRegexFinalB="zcat /data/log/2021-"+str(LastDateRegex)+"-[0-2][0-9]*/"+str(FIP)+ ''' -traffic* | grep "$p" | grep -E '''
    endStringB=''' | awk -F "=" '{print $11 "\t" $3 "\t" $12 "\t" $4 "\t" $10 "\t" $19 "\t" $23 "\t" $29 "\t" $30 "\t" $35}' |  awk  -F"ankitdeshwal" '{a[$1]++}END{for(i in a) {print i, a[i]}}' | sed 's/cs5//g;s/deviceInboundInterface//g;s/dst//g;s/sourceTranslatedAddress//g;s/cs4//g;s/sourceTranslatedPort//g;s/PanOSPacketsSent//g;s/cn3//g;s/cat//g;s/act//g'>> ''' +str(VLN)+'''-broader"$p".txt'''
    broader=dateRegexFinalB+IPFinalRegex+endStringB
    return toptalker


def regCreatorB(daterange,FIP,Sub,VLN,wsNet,wsDate,subsplitNet):
    FOct = subsplitNet.split('.')[0]
    StOct = subsplitNet.split('.')[1]
    TOct = subsplitNet.split('.')[2]
    LOct=subsplitNet.split('.')[3]
    IPRegx=FOct+"\."+StOct+"\."+TOct+"\."
    LastRegx=""
    print("ankit")
    print(str(LOct))
    for i in range(1, wsNet.max_row + 1):
        print(str(wsNet.cell(row=i, column=1).value))
        if wsNet.cell(row=i, column=1).value is not None:
            print("inside row "+str(wsNet.cell(row=i, column=1).value))
            if str(wsNet.cell(row=i, column=1).value) == str(LOct):
                for j in range(1, wsNet.max_column  + 1):
                    print("inside column " + str(wsNet.cell(row=i, column=1).value))
                    if wsNet.cell(row=i, column=j).value is not None:
                        if j!=1:
                            LastRegx=LastRegx+IPRegx+str(wsNet.cell(row=i, column=j).value)+"\s|"
    for i in range(1, wsDate.max_row + 1):
        if wsDate.cell(row=i, column=3).value is not None:
            if str(wsDate.cell(row=i, column=3).value) == str(daterange):
                LastDateRegex=wsDate.cell(row=i, column=4).value

    IPFinalRegex="'"+LastRegx[:-1]+"'"
    dateRegexFinalB="zcat /data/log/2021-"+str(LastDateRegex)+"-[0-2][0-9]*/"+str(FIP)+ '''-traffic* | grep "$p" | grep -E '''
    endStringB=''' | awk -F "=" '{print $11 "\t" $3 "\t" $12 "\t" $4 "\t" $10 "\t" $19 "\t" $23 "\t" $29 "\t" $30 "\t" $35}' |  awk  -F"ankitdeshwal" '{a[$1]++}END{for(i in a) {print i, a[i]}}' | sed 's/cs5//g;s/deviceInboundInterface//g;s/dst//g;s/sourceTranslatedAddress//g;s/cs4//g;s/sourceTranslatedPort//g;s/PanOSPacketsSent//g;s/cn3//g;s/cat//g;s/act//g'>> ''' +str(VLN)+'''-broader"$p".txt'''
    broader=dateRegexFinalB+IPFinalRegex+endStringB
    return broader

def myClick(daterange,FIP,Sub,VLN,Scripttype):
    wb = openpyxl.load_workbook("REGEX.xlsx")
    sheetNameList = wb.sheetnames
    subsplit=Sub.split('/')
    subsplitMask=subsplit[1]
    subsplitNet = subsplit[0]
    sheetName=subsplitMask
    wsNet=wb[sheetName]
    wsDate=wb[sheetNameList[7]]

    if Scripttype == "TopTalker":
        value= regCreatorT(daterange,FIP,Sub,VLN,wsNet,wsDate,subsplitNet)
        messageBOX = Text(root, width=100, height=10, bg="yellow", fg="black", borderwidth=5)
        messageBOX.grid(row=6, columnspan=3)
        messageBOX.insert(0.0, value)
        print("insideTop")
    if Scripttype=="Broader":
        value=regCreatorB(daterange, FIP, Sub, VLN,wsNet,wsDate,subsplitNet)
        messageBOX = Text(root, width=100, height=10, bg="yellow", fg="black", borderwidth=5)
        messageBOX.grid(row=6, columnspan=3)
        TOTAL="while read p; do \n" + value +"\ndone < multisourcedest.txt"
        messageBOX.insert(0.0, TOTAL)



MR=["Jan-Feb","Feb-Mar","Mar-Apr","Apr-May","May-Jun","Jun-Jul","Jul-Aug","Aug-Sep","Sep-Oct","Oct-Nov","Nov-Dec","Jan-Feb-Mar","Feb-Mar-Apr","Mar-Apr-May","Apr-May-Jun","May-Jun-Jul","Jun-Jul-Aug","Jul-Aug-Sep","Aug-Sep-Oct","Sep-Oct-Nov","Oct-Nov-Dec","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
RB=["TopTalker","Broader"]
root=Tk()
root.geometry("1000x500")
variable = StringVar(root)
variable.set(MR[0])
rbvariable=StringVar()
rbvariable.set(RB[0])

#elements
DDLabel=Label(root,text="Month Range : ")
DropDown= OptionMenu(root, variable, *MR)

FWIPLabel=Label(root,text="Firewall ip : ")
FWIP=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)


subnetLabel=Label(root,text="subnet/mask : ")
subnet=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)

VlanLabel=Label(root,text="Vlan Number : ")
Vlan=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)
Radiobutton(root,text="TopTalker ",variable=rbvariable,value=RB[0]).grid(row=4,column=0)
Radiobutton(root,text="Broader ",variable=rbvariable,value=RB[1]).grid(row=4,column=1)

myButton = Button(root,text="Generate Regex",command=lambda: myClick(variable.get(),FWIP.get(),subnet.get(),Vlan.get(),rbvariable.get()),fg="yellow",bg="black")
myButton.grid(row=5,column=0)

#Location
DDLabel.grid(row=0,column=0)
DropDown.grid(row=0,column=1)
FWIPLabel.grid(row=1,column=0)
FWIP.grid(row=1,column=1)
subnetLabel.grid(row=2,column=0)
subnet.grid(row=2,column=1)
VlanLabel.grid(row=3,column=0)
Vlan.grid(row=3,column=1)

root.mainloop()