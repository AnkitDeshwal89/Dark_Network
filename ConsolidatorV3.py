import openpyxl
from tkinter import *
from tkinter import filedialog
root = Tk()
root.geometry('700x200')
def browseFiles():
    filename = filedialog.askopenfilename(initialdir="c:", title="Select a File",
                                          filetypes=(("Excel File", "*.xlsx"), ("all files", "*.*")))
    wbn.insert(0,filename)



def zonerename(sheet, wbname, wb,zone):
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=9).value is not None:
            datac5 = sheet.cell(row=i, column=5).value
            datac9 = sheet.cell(row=i, column=9).value
            datac5split = datac5.split(';')
            datac9split = datac9.split(';')
            if zone in datac5split:
                if len(datac5split) > 1 and zone not in datac9split:
                    print("col 5 {}".format(datac5split))
                    sheet.cell(row=i, column=5).value = zone

            if zone in datac9split:
                if len(datac9split) > 1 and zone not in datac5split:
                    print("col 9 {}".format(datac9split))
                    sheet.cell(row=i, column=9).value = zone
    wb.save(wbname)


def dup_identify_light(sheet, wbname, wb):
    for i in range(1, sheet.max_row + 1):
        dupFirstTime = True
        if sheet.cell(row=i, column=9).value is not None:
            srczn = sheet.cell(row=i, column=5).value
            dstzn = sheet.cell(row=i, column=9).value
            srcip = sheet.cell(row=i, column=6).value
            dstip = sheet.cell(row=i, column=10).value
            app = sheet.cell(row=i, column=14).value
            service = sheet.cell(row=i, column=15).value
            for j in range(i + 1, sheet.max_row + 1):
                srcznnxt = sheet.cell(row=j, column=5).value
                dstznnxt = sheet.cell(row=j, column=9).value
                srcipnxt = sheet.cell(row=j, column=6).value
                dstipnxt = sheet.cell(row=j, column=10).value
                appnxt = sheet.cell(row=j, column=14).value
                servicenxt = sheet.cell(row=j, column=15).value
                if srczn == srcznnxt and dstzn == dstznnxt and srcip == srcipnxt and dstip == dstipnxt and app == appnxt and servicenxt == service:
                    if sheet.cell(row=i, column=21).value != "duplicate":
                        sheet.cell(row=j, column=21).value = "duplicate"
                        sheet.cell(row=j, column=22).value = i
                        print("dup addeed {}".format(sheet.cell(row=j, column=21).value))
                        if dupFirstTime:
                            sheet.cell(row=i, column=21).value = "duplicate"
                            sheet.cell(row=i, column=22).value = i
                            sheet.cell(row=i, column=23).value = str(sheet.cell(row=i, column=2).value)
                            dupFirstTime = False
                        sheet.cell(row=i, column=23).value = sheet.cell(row=i, column=23).value + " , "+ str(sheet.cell(row=j, column=2).value)
    wb.save(wbname)


def dup_remover(sheet, wbname, wb):
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=9).value is not None:
            if sheet.cell(row=i, column=21).value == "duplicate":
                for j in range(i + 1, sheet.max_row + 1):
                    if sheet.cell(row=i, column=22).value == sheet.cell(row=j, column=22).value:
                        sheet.delete_rows(j)
    wb.save(wbname)


def zonesorting(sheet, ws4, ws5, wbname, wb):
    for i in range(1, sheet.max_row + 1):
        print("Starting merger = {} ".format(sheet.cell(row=i, column=2).value))
        AddFirstTime = True
        if sheet.cell(row=i, column=9).value is not None:
            srczn = sheet.cell(row=i, column=5).value
            dstzn = sheet.cell(row=i, column=9).value
            for j in range(i + 1, sheet.max_row + 1):
                srcznnxt = sheet.cell(row=j, column=5).value
                dstznnxt = sheet.cell(row=j, column=9).value
                if srczn == srcznnxt and dstzn == dstznnxt:
                    if (str(sheet.cell(row=j, column=2).value) + "," + sheet.cell(row=j, column=19).value) not in diclst:
                        diclst.append(str(sheet.cell(row=j, column=2).value)+ "," + sheet.cell(row=j,
                                                                                           column=19).value)  # for condition checking
                        if AddFirstTime == True:
                            diclst.append(str(sheet.cell(row=i, column=2).value) + "," + sheet.cell(row=i,
                                                                                               column=19).value)  # for condition checking
                            sqi = sheet.cell(row=i, column=1).value
                            Namei = str(sheet.cell(row=i, column=2).value)
                            Tagsi = sheet.cell(row=i, column=3).value
                            Typei = sheet.cell(row=i, column=4).value
                            Source_Zonei = sheet.cell(row=i, column=5).value
                            Source_Addressi = sheet.cell(row=i, column=6).value
                            Source_Useri = sheet.cell(row=i, column=7).value
                            Source_HIP_Profilei = sheet.cell(row=i, column=8).value
                            Destination_Zonei = sheet.cell(row=i, column=9).value
                            Destination_Addressi = sheet.cell(row=i, column=10).value
                            Rule_Usage_Hit_Counti = sheet.cell(row=i, column=11).value
                            Rule_Usage_Last_Hiti = sheet.cell(row=i, column=12).value
                            Rule_Usage_First_Hiti = sheet.cell(row=i, column=13).value
                            Application_Service_Actioni = sheet.cell(row=i, column=14).value
                            Servicei = sheet.cell(row=i, column=15).value
                            Actioni = sheet.cell(row=i, column=16).value
                            Profile_Optionsi = sheet.cell(row=i, column=17).value
                            Optionsi = sheet.cell(row=i, column=18).value
                            Firewalli = sheet.cell(row=i, column=19).value
                            diclstsheet.append([sqi, Namei, Tagsi, Typei, Source_Zonei, Source_Addressi, Source_Useri,
                                                Source_HIP_Profilei, Destination_Zonei, Destination_Addressi,
                                                Rule_Usage_Hit_Counti, Rule_Usage_Last_Hiti, Rule_Usage_First_Hiti,
                                                Application_Service_Actioni, Servicei, Actioni, Profile_Optionsi,
                                                Optionsi, Firewalli])  # for adding new sheet
                            print("adding merger first = {} ".format(sheet.cell(row=i, column=2).value))
                            sheet.cell(row=i, column=20).value = "added"
                            AddFirstTime = False

                        sq = sheet.cell(row=j, column=1).value
                        Name = str(sheet.cell(row=j, column=2).value)
                        Tags = sheet.cell(row=j, column=3).value
                        Type = sheet.cell(row=j, column=4).value
                        Source_Zone = sheet.cell(row=j, column=5).value
                        Source_Address = sheet.cell(row=j, column=6).value
                        Source_User = sheet.cell(row=j, column=7).value
                        Source_HIP_Profile = sheet.cell(row=j, column=8).value
                        Destination_Zone = sheet.cell(row=j, column=9).value
                        Destination_Address = sheet.cell(row=j, column=10).value
                        Rule_Usage_Hit_Count = sheet.cell(row=j, column=11).value
                        Rule_Usage_Last_Hit = sheet.cell(row=j, column=12).value
                        Rule_Usage_First_Hit = sheet.cell(row=j, column=13).value
                        Application_Service_Action = sheet.cell(row=j, column=14).value
                        Service = sheet.cell(row=j, column=15).value
                        Action = sheet.cell(row=j, column=16).value
                        Profile_Options = sheet.cell(row=j, column=17).value
                        Options = sheet.cell(row=j, column=18).value
                        Firewall = sheet.cell(row=j, column=19).value
                        diclstsheet.append(
                            [sq, Name, Tags, Type, Source_Zone, Source_Address, Source_User, Source_HIP_Profile,
                             Destination_Zone, Destination_Address, Rule_Usage_Hit_Count, Rule_Usage_Last_Hit,
                             Rule_Usage_First_Hit, Application_Service_Action, Service, Action, Profile_Options,
                             Options, Firewall])
                        print("adding merger notfirst = {} ".format(sheet.cell(row=j, column=2).value))
                        sheet.cell(row=j, column=20).value = "added"
    for i in range(0, len(diclstsheet)):
        for j in range(0, len(diclstsheet[i])):
            ws4.cell(row=i + 1, column=j + 1).value = diclstsheet[i][j]
    wb.save(wbname)

    print("===========================================================================================")
    for item in range(1, sheet.max_row + 1):
        print("Starting unqiue = {} ".format(sheet.cell(row=item, column=2).value))
        if sheet.cell(row=item, column=9).value is not None:
            if (str(sheet.cell(row=item, column=2).value) + "," + sheet.cell(row=item, column=19).value) not in diclst:
                dicitem.append(str(sheet.cell(row=item, column=2).value) + "," + sheet.cell(row=item, column=19).value)
                psrc = sheet.cell(row=item, column=5).value
                pdst = sheet.cell(row=item, column=9).value
                pru = str(sheet.cell(row=item, column=2).value)
                print("{},{},{}".format(pru, psrc, pdst))
                sq = sheet.cell(row=item, column=1).value
                Name = str(sheet.cell(row=item, column=2).value)
                Tags = sheet.cell(row=item, column=3).value
                Type = sheet.cell(row=item, column=4).value
                Source_Zone = sheet.cell(row=item, column=5).value
                Source_Address = sheet.cell(row=item, column=6).value
                Source_User = sheet.cell(row=item, column=7).value
                Source_HIP_Profile = sheet.cell(row=item, column=8).value
                Destination_Zone = sheet.cell(row=item, column=9).value
                Destination_Address = sheet.cell(row=item, column=10).value
                Rule_Usage_Hit_Count = sheet.cell(row=item, column=11).value
                Rule_Usage_Last_Hit = sheet.cell(row=item, column=12).value
                Rule_Usage_First_Hit = sheet.cell(row=item, column=13).value
                Application_Service_Action = sheet.cell(row=item, column=14).value
                Service = sheet.cell(row=item, column=15).value
                Action = sheet.cell(row=item, column=16).value
                Profile_Options = sheet.cell(row=item, column=17).value
                Options = sheet.cell(row=item, column=18).value
                Firewall = sheet.cell(row=item, column=19).value

                dicitemsheet.append([sq, Name, Tags, Type, Source_Zone, Source_Address, Source_User, Source_HIP_Profile,
                                     Destination_Zone, Destination_Address, Rule_Usage_Hit_Count, Rule_Usage_Last_Hit,
                                     Rule_Usage_First_Hit, Application_Service_Action, Service, Action, Profile_Options,
                                     Options, Firewall])
                print("adding unqiue = {} ".format(sheet.cell(row=item, column=2).value))
                sheet.cell(row=item, column=20).value = "added"
    for i in range(0, len(dicitemsheet)):
        for j in range(0, len(dicitemsheet[i])):
            ws5.cell(row=i + 1, column=j + 1).value = dicitemsheet[i][j]
    wb.save(wbname)
    count = len(diclst) + len(dicitem)
    print("total item {} ".format(count))
    print(len(diclst))
    print(len(dicitem))


def workbookcleaner(wbname, fullrun='yes'):
    if fullrun == 'yes':
        wb = openpyxl.load_workbook(wbname)
        #sheet remove start
        sheetremovelist = wb.sheetnames
        for i in range(len(sheetremovelist)-1):
            wb.remove(wb[sheetremovelist[i+1]])
        # sheet remove ends and sheet addtion starts below
        ws2 = wb.create_sheet("NoZone", 1)
        ws3 = wb.create_sheet("Required", 2)
        ws4 = wb.create_sheet("MergerPolicy", 3)
        ws5 = wb.create_sheet("UniquePolicy", 4)
        sheetNameList = wb.sheetnames
        sheetName = sheetNameList[0]
        ws1 = wb[sheetName]
        ws1.cell(row=1, column=1).value = "Sno"
        rowlist=['Sno', 'Name', 'Tags', 'Type', 'Source Zone', 'Source Address', 'Source User', 'Source HIP Profile', 'Destination Zone', 'Destination Address', 'Rule Usage Hit Count', 'Rule Usage Last Hit', 'Rule Usage First Hit', 'Application', 'Service', 'Action', 'Profile', 'Options']
        for i in range(1, ws1.max_column + 1):
             if(ws1.cell(row=1, column=i).value)==rowlist[i-1]:
                 print("Good to proceed")
             else:
                 print("check your sheet  columns ")
                 sys.exit()
        print(rowlist)
        wb.save(wbname)
        return [ws1, ws2, ws3, ws4, ws5, wb]
    else:
        wb = openpyxl.load_workbook(wbname)
        return wb


def selectrequired(sheet, ws2, ws3, wbname, wb,zone):
    global r, c
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=9).value is not None:
            datac5 = sheet.cell(row=i, column=5).value
            datac9 = sheet.cell(row=i, column=9).value
            datac5split = datac5.split(';')
            datac9split = datac9.split(';')

            if zone not in datac9split and zone not in datac5split:
                if zoneany not in datac9split and zoneany not in datac5split:
                    for j in range(1, sheet.max_column + 1):
                        cell1 = sheet.cell(row=i, column=j).value
                        ws2.cell(row=i, column=j).value = cell1
                    #                print(sheet.cell(row=i,column=1).value)
                    c += 1
            if zone in datac9split or zone in datac5split or zoneany in datac9split or zoneany in datac5split:
                for j in range(1, sheet.max_column + 1):
                    cell1 = sheet.cell(row=i, column=j).value
                    ws3.cell(row=i, column=j).value = cell1
                #                print(sheet.cell(row=i,column=1).value)
                r += 1
    print("require = {} , Not require = {}".format(r, c))
    wb.save(wbname)


def writefwname(fwname, ws3, wbname, wb, twfwmerger):
    if twfwmerger == 'yes':
        pass
    else:
        for i in range(1, ws3.max_row + 1):
            if ws3.cell(row=i, column=9).value is not None:
                ws3.cell(row=i, column=19).value = fwname
    wb.save(wbname)

deletelstdeny = []
deletelstdeny2 = []

def deletedeny(sheet,sheet2 ,wbname, wb):
    countm = 0
    countu = 0
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=16).value == 'Deny' or sheet.cell(row=i, column=16).value == 'Drop' or sheet.cell(row=i, column=16).value == 'deny' or sheet.cell(row=i, column=16).value == 'drop':
            countm = countm +1
    for i in range(1, sheet2.max_row + 1):
        if sheet2.cell(row=i, column=16).value == 'Deny' or sheet2.cell(row=i, column=16).value == 'Drop' or sheet2.cell(row=i, column=16).value == 'deny' or sheet2.cell(row=i, column=16).value == 'drop':
            countu = countu + 1

    for j in range(0,countm):
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=16).value == 'Deny' or sheet.cell(row=i, column=16).value == 'Drop' or sheet.cell(row=i, column=16).value == 'deny' or sheet.cell(row=i, column=16).value == 'drop':
                sheet.delete_rows(i)

    for j in range(0,countu):
        for i in range(1, sheet2.max_row + 1):
            if sheet2.cell(row=i, column=16).value == 'Deny' or sheet2.cell(row=i, column=16).value == 'Drop' or sheet2.cell(row=i, column=16).value == 'deny' or sheet2.cell(row=i, column=16).value == 'drop':
                sheet2.delete_rows(i)
    wb.save(wbname)

deleterow =[]
def merger(sheet,wbname, wb):
#    wb = workbookcleaner(wbname, fullrun)
#    sheet = wb[wb.sheetnames[3]]
    ws6 = wb.create_sheet('MergerPolicy_Backup', 6)
    for i in range(1, sheet.max_row + 1):
        for j in range(1,sheet.max_column +1):
            ws6.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value
    countmerger=0
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=9).value != '0':
            tmpdsti = ''
            tmpsrci = ''
            merFirstTime = True
            if sheet.cell(row=i, column=9).value is not None:
                srczn = sheet.cell(row=i, column=5).value
                dstzn = sheet.cell(row=i, column=9).value
                srcip = sheet.cell(row=i, column=6).value
                dstip = sheet.cell(row=i, column=10).value
                app = sheet.cell(row=i, column=14).value
                service = sheet.cell(row=i, column=15).value
                for j in range(i + 1, sheet.max_row + 1):
                    srcznnxt = sheet.cell(row=j, column=5).value
                    dstznnxt = sheet.cell(row=j, column=9).value
                    srcipnxt = sheet.cell(row=j, column=6).value
                    dstipnxt = sheet.cell(row=j, column=10).value
                    appnxt = sheet.cell(row=j, column=14).value
                    servicenxt = sheet.cell(row=j, column=15).value
                    if srczn == srcznnxt and dstzn == dstznnxt and app == appnxt and servicenxt == service:
                        if srcip != srcipnxt and dstip == dstipnxt and dstip != zoneany:
                            if (str(sheet.cell(row=j, column=2).value) + "," + sheet.cell(row=j,column=19).value) not in mergdic:
                                mergdic.append(str(sheet.cell(row=j, column=2).value) + "," + sheet.cell(row=j, column=19).value)
                                if merFirstTime:
                                    mergdic.append(str(sheet.cell(row=i, column=2).value) + "," + sheet.cell(row=i, column=19).value)
                                    tmpsrci = srcip
                                    tmpponamedst = str(sheet.cell(row=i, column=2).value) + "." + sheet.cell(row=i, column=19).value
                                    merFirstTime = False
                                print(str(sheet.cell(row=j, column=2).value) + " Same Dest  can be merged with " + sheet.cell(row=i,column=2).value)
                                tmpsrci = tmpsrci + ',' + srcipnxt
                                tmpponamedst = tmpponamedst + ',' + str(sheet.cell(row=j, column=2).value) + "." + sheet.cell(row=j, column=19).value
                                sheet.cell(row=j, column=20).value = 'mergerdelete'
                                countmerger = countmerger +1

                        if srcip == srcipnxt and srcip != zoneany and dstip != dstipnxt:
                            if (str(sheet.cell(row=j, column=2).value) + "," + sheet.cell(row=j,column=19).value) not in mergdic:
                                mergdic.append(str(sheet.cell(row=j, column=2).value) + "," + sheet.cell(row=j, column=19).value)
                                if merFirstTime:
                                    mergdic.append(str(sheet.cell(row=i, column=2).value) + "," + sheet.cell(row=i, column=19).value)
                                    tmpdsti = dstip
                                    tmpponamesrc = str(sheet.cell(row=i, column=2).value) + "." + sheet.cell(row=i, column=19).value
                                    merFirstTime = False
                                print(str(sheet.cell(row=j, column=2).value )+ " Same Source can be merged with " + sheet.cell(row=i,column=2).value)
                                tmpdsti = tmpdsti + ',' + dstipnxt
                                tmpponamesrc = tmpponamesrc + ',' + str(sheet.cell(row=j, column=2).value) + "." + sheet.cell(row=j, column=19).value
                                sheet.cell(row=j, column=20).value = 'mergerdelete'
                                countmerger = countmerger +1
                if len(tmpdsti) > 0:
                    print(tmpdsti)
                    sheet.cell(row=i, column=10).value =  tmpdsti
                    sheet.cell(row=i,column=2).value = tmpponamesrc
                if len(tmpsrci) > 0:
                    print(tmpsrci)
                    sheet.cell(row=i, column=6).value =  tmpsrci
                    sheet.cell(row=i,column=2).value = tmpponamedst
    print("Total Policy that will be merged = {}".format(countmerger))
    for i in range(0,countmerger):
        for j in range(1,sheet.max_row +1):
            if sheet.cell(row=j,column=20).value =='mergerdelete':
                print("deleted merger {}".format(sheet.cell(row=j, column=2).value))
                sheet.delete_rows(j)

    wb.save(wbname)


def Finalmerger(ws4, ws5, wbname, wb,ws1):
    count = 0
    ws6 =  wb.create_sheet('Final_Merger',6)
    print("merger max row {}".format(ws4.max_row +1))
    print("unqiue max row {}".format(ws5.max_row + 1))
    for i in range(1,ws4.max_row +1):
        if ws4.cell(row=i,column=9).value is not None:
            count += 1
            print(count)
            for j in range(1,ws4.max_column +1):
                ws6.cell(row=i,column=j).value = ws4.cell(row=i,column=j).value


    print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    for i in range(1,ws5.max_row +1):
        if ws5.cell(row=i,column=9).value is not None:
            print(count + i)
            for j in range(1,ws5.max_column +1):
                ws6.cell(row=count+i,column=j).value =ws5.cell(row=i,column=j).value
    print("merger done")
    print("inserting row")
    # insert 1st row , 3 colums consiting of syslog,comment,engineer
    ws6.insert_rows(1,1)
    for i in range(1,ws1.max_column +1):
        ws6.cell(row=1, column=i).value = ws1.cell(row=1, column=i).value
    ws6.insert_cols(1,3)
    ws6.cell(row=1, column=1).value = "syslog"
    ws6.cell(row=1, column=2).value = "engineer"
    ws6.cell(row=1, column=3).value = "comment"
    ws7 = wb.create_sheet('Syslog', 7)
    wb.save(wbname)

#fetch data from firewall downloaded sheet
def dataFetcher(rownumber, wbname,fw_tagssplit,zone,zoneany):
    wb = openpyxl.load_workbook(wbname)
    ws2 = wb.create_sheet("DI", 1)
    sheetNameList = wb.sheetnames
    sheetName = sheetNameList[0]
    sheet = wb[sheetName]
    rownumber.append(1)
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=9).value is not None:
            datac6 = sheet.cell(row=i, column=6).value
            datac10 = sheet.cell(row=i, column=10).value
            datac6split = datac6.split(';')
            datac10split = datac10.split(';')
            datac5 = sheet.cell(row=i, column=5).value
            datac9 = sheet.cell(row=i, column=9).value
            datac5split = datac5.split(';')
            datac9split = datac9.split(';')
            for j in datac6split:
                for k in fw_tagssplit:
                    if j.lower().find(k.lower()) != -1:
                        rownumber.append(i)
            for j in datac10split:
                for k in fw_tagssplit:
                    if j.lower().find(k.lower()) != -1:
                        rownumber.append(i)

            if zone in datac9split or zone in datac5split or zoneany in datac9split or zoneany in datac5split or ipany in datac6split or ipany in datac10split:
                rownumber.append(i)

    rownumber = list(set(rownumber))

    for i in range(len(rownumber)):
        for j in range(1, sheet.max_column + 1):
            cell1 = sheet.cell(row=rownumber[i], column=j).value
            ws2.cell(row=i + 1, column=j).value = cell1

    wb.remove(wb[sheetNameList[0]])
    wb.save(wbname)


def myClick(rownumber,fullrun,zoneany):
    workbookname = wbn.get().strip('')
    zone = zn.get().strip('')
    fw_tags = subtags.get().split(";")
    firewallname = fwname.get().strip('')
    #inputlist.append(workbookname,zone,fw_tags,firewallname)
    dataFetcher(rownumber, workbookname,fw_tags,zone,zoneany)
    wsheetlist = workbookcleaner(workbookname, fullrun)
    wb = wsheetlist[5]
    selectrequired(wsheetlist[0], wsheetlist[1], wsheetlist[2], workbookname, wb,zone)
    writefwname(firewallname, wsheetlist[2], workbookname, wb, twfwmerger)
    zonerename(wsheetlist[2], workbookname, wb,zone)
    zonesorting(wsheetlist[2], wsheetlist[3], wsheetlist[4], workbookname, wb)
    dup_identify_light(wsheetlist[3], workbookname, wb)
    dup_remover(wsheetlist[3], workbookname, wb)
    merger(wsheetlist[3], workbookname, wb)
    deletedeny(wsheetlist[3], wsheetlist[4], workbookname, wb)
    Finalmerger(wsheetlist[3], wsheetlist[4], workbookname, wb, wsheetlist[0])
    myLable=Label(root,text="**********Completed************")
    myLable.grid(row=5,column=1)


mergdic = []
zoneany = 'any'
c = 0
r = 0
diclst = []
diclstsheet = []
dicitem = []
dicitemsheet = []
duplst = []
wb = ''
fullrun = "yes" #input('put  yes for now : ')
twfwmerger = "no" #input('yes or no if merger of two sheet like SESA and SELI VLAN 88 : ')
rownumber=[]
ipany = "any"
wsheetlist = []

button_explore = Button(root,text = "Browse Files",command =browseFiles)
wbnlabel=Label(root,text="WorkBookName : ")
wbn=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)

znlabel=Label(root,text="Zone Name : ")
zn=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)

taglabel=Label(root,text="Subnet Tags : ")
subtags=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)

fwnamelabel=Label(root,text="Firewall Name : ")
fwname=Entry(root,width=50,bg="yellow",fg="black",borderwidth=5)


#display

wbnlabel.grid(row=0,column=0)
wbn.grid(row=0,column=1)
button_explore.grid(row=0,column=3)
znlabel.grid(row=1,column=0)
zn.grid(row=1,column=1)
taglabel.grid(row=2,column=0)
subtags.grid(row=2,column=1)
fwnamelabel.grid(row=3,column=0)
fwname.grid(row=3,column=1)
#variables
inputlist=[]
myButton = Button(root,text="Run Consolidator",command=lambda: myClick(rownumber,fullrun,zoneany),fg="yellow",bg="black")
myButton.grid(row=4,column=0)
root.mainloop()







